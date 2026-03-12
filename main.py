import os
import re
import asyncio
import logging
from typing import List, Dict, Optional

import requests
from telegram import Update
from telegram.ext import (
    Application,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

TG_TOKEN = os.getenv("TG_BOT_TOKEN")

SELLERS = [
    {"name": "МС", "url": "https://www.wildberries.ru/seller/92351?sort=newly&page=1"},
    {"name": "ЦР", "url": "https://www.wildberries.ru/seller/870386?sort=newly&page=1"},
]

LIMIT = 20
OUTPUT_XLSX = "wb_products.xlsx"
TEMP_IMG_DIR = "temp_images"
BOT_USERNAME = "vatras_bot"

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("wb_bot")


def extract_nm_id(url: str) -> Optional[int]:
    m = re.search(r"/catalog/(\d+)/detail\.aspx", url)
    if m:
        return int(m.group(1))
    return None


def download_image(image_url: str, file_path: str) -> Optional[str]:
    try:
        response = requests.get(image_url, timeout=30)
        response.raise_for_status()
        with open(file_path, "wb") as f:
            f.write(response.content)
        return file_path
    except Exception as e:
        logger.warning("Не удалось скачать картинку %s: %s", image_url, e)
        return None


def prepare_excel_image(image_path: str, max_width: int = 160, max_height: int = 160) -> Optional[str]:
    try:
        img = PILImage.open(image_path)
        img.thumbnail((max_width, max_height))
        prepared_path = image_path.rsplit(".", 1)[0] + "_prepared.png"
        img.save(prepared_path, format="PNG")
        return prepared_path
    except Exception as e:
        logger.warning("Не удалось подготовить картинку %s: %s", image_path, e)
        return None


def format_sheet(ws):
    headers = ["Наименование", "Категория", "Картинка", "Ссылка"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 60


def fill_sheet(ws, items: List[Dict], sheet_prefix: str):
    current_row = 2

    for idx, item in enumerate(items, start=1):
        ws.cell(row=current_row, column=1, value=item.get("title", ""))
        ws.cell(row=current_row, column=2, value=item.get("category", ""))

        link = item.get("url", "")
        link_cell = ws.cell(row=current_row, column=4, value=link)
        if link:
            link_cell.hyperlink = link
            link_cell.style = "Hyperlink"

        ws.cell(row=current_row, column=1).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=current_row, column=2).alignment = Alignment(vertical="top", wrap_text=True)

        image_url = item.get("image")
        if image_url:
            raw_path = os.path.join(TEMP_IMG_DIR, f"{sheet_prefix}_img_{idx}.jpg")
            downloaded = download_image(image_url, raw_path)

            if downloaded:
                prepared = prepare_excel_image(downloaded)
                if prepared:
                    try:
                        img = XLImage(prepared)
                        ws.add_image(img, f"C{current_row}")
                    except Exception as e:
                        logger.warning("Не удалось вставить картинку в Excel: %s", e)

        ws.row_dimensions[current_row].height = 130
        current_row += 1


def save_to_xlsx(sellers_data: Dict[str, List[Dict]], path: str) -> str:
    os.makedirs(TEMP_IMG_DIR, exist_ok=True)

    wb = Workbook()
    first_sheet = True

    for sheet_name, items in sellers_data.items():
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        format_sheet(ws)
        fill_sheet(ws, items, sheet_name)

    wb.save(path)
    return path


async def collect_product_links(page, limit: int) -> List[str]:
    selectors = [
        "a.product-card__link",
        "article a[href*='/catalog/'][href*='/detail.aspx']",
        "a[href*='/catalog/'][href*='/detail.aspx']",
    ]

    found_selector = None
    for selector in selectors:
        try:
            await page.wait_for_selector(selector, timeout=10000)
            found_selector = selector
            logger.info("Найден селектор: %s", selector)
            break
        except PlaywrightTimeoutError:
            logger.info("Селектор не найден: %s", selector)

    if not found_selector:
        raise RuntimeError("На странице продавца не найдены карточки товаров")

    initial_links = await page.locator("a[href*='/catalog/'][href*='/detail.aspx']").evaluate_all(
        "elements => elements.map(el => el.href).filter(Boolean)"
    )

    collected_links = list(initial_links)

    for _ in range(5):
        await page.mouse.wheel(0, 2800)
        await page.wait_for_timeout(1200)

        new_links = await page.locator("a[href*='/catalog/'][href*='/detail.aspx']").evaluate_all(
            "elements => elements.map(el => el.href).filter(Boolean)"
        )
        collected_links.extend(new_links)

        if len(collected_links) >= limit * 3:
            break

    result = []
    seen = set()

    for href in collected_links:
        nm_id = extract_nm_id(href)
        if not nm_id:
            continue
        if nm_id in seen:
            continue

        seen.add(nm_id)
        result.append(href)

        if len(result) >= limit:
            break

    logger.info("Собрано ссылок: %s", len(result))
    return result


async def get_text_by_selectors(page, selectors: List[str]) -> Optional[str]:
    for selector in selectors:
        try:
            locator = page.locator(selector).first
            if await locator.count() > 0:
                text = await locator.inner_text()
                text = " ".join(text.split())
                if text:
                    return text
        except Exception:
            continue
    return None


async def get_image_by_selectors(page, selectors: List[str]) -> Optional[str]:
    for selector in selectors:
        try:
            locator = page.locator(selector).first
            if await locator.count() > 0:
                src = await locator.get_attribute("src")
                current_src = await locator.get_attribute("currentSrc")
                candidate = current_src or src
                if candidate:
                    if candidate.startswith("//"):
                        candidate = "https:" + candidate
                    elif candidate.startswith("/"):
                        candidate = "https://www.wildberries.ru" + candidate
                    return candidate
        except Exception:
            continue
    return None


async def parse_product_page(context, url: str) -> Dict:
    page = await context.new_page()
    page.set_default_timeout(10000)

    try:
        logger.info("Открываю карточку: %s", url)
        await page.goto(url, wait_until="domcontentloaded", timeout=20000)
        await page.wait_for_timeout(2000)

        title = None
        try:
            await page.wait_for_selector("h3", timeout=4000)
            title = await page.locator("h3").first.inner_text()
            title = " ".join(title.split())
        except Exception:
            pass

        if not title:
            title = await get_text_by_selectors(page, [
                "h3",
                ".product-page h3",
                "[class*='product'] h3",
            ])

        if not title:
            try:
                html = await page.content()
                match = re.search(r"<h3[^>]*>(.*?)</h3>", html, re.S)
                if match:
                    title = re.sub(r"<.*?>", "", match.group(1)).strip()
                    title = " ".join(title.split())
            except Exception:
                pass

        category = None
        try:
            await page.wait_for_selector("span.categoryLinkCategory--VSJ8c", timeout=4000)
            category = await page.locator("span.categoryLinkCategory--VSJ8c").first.inner_text()
            category = " ".join(category.split())
        except Exception:
            pass

        if not category:
            category = await get_text_by_selectors(page, [
                "span.categoryLinkCategory--VSJ8c",
                "span[class*='categoryLinkCategory']",
            ])

        if not category:
            try:
                html = await page.content()
                match = re.search(r'class="categoryLinkCategory--VSJ8c">(.*?)</span>', html, re.S)
                if match:
                    category = re.sub(r"<.*?>", "", match.group(1)).strip()
                    category = " ".join(category.split())
            except Exception:
                pass

        image = await get_image_by_selectors(page, [
            ".swiper-slide-active img",
            ".product-page__slider img",
            ".photo-zoom__preview img",
            ".j-image-container img",
            "img[src*='wbbasket']",
            "img",
        ])

        nm_id = extract_nm_id(url)

        if not title:
            title = f"Товар {nm_id}" if nm_id else "Без названия"

        if not category:
            category = "Категория не найдена"

        if not image and nm_id:
            vol = nm_id // 100000
            part = nm_id // 1000
            image = f"https://basket-01.wbbasket.ru/vol{vol}/part{part}/{nm_id}/images/big/1.webp"

        return {
            "title": title,
            "category": category,
            "image": image,
            "url": url,
            "nm_id": nm_id,
        }

    finally:
        await page.close()


async def safe_parse_product_page(context, url: str, idx: int) -> Optional[Dict]:
    try:
        item = await asyncio.wait_for(parse_product_page(context, url), timeout=20)

        if idx == 1 and (
            item.get("category") == "Категория не найдена"
            or item.get("title", "").startswith("Товар ")
        ):
            await asyncio.sleep(1.5)
            item = await asyncio.wait_for(parse_product_page(context, url), timeout=20)

        return item

    except asyncio.TimeoutError:
        logger.warning("Таймаут на карточке: %s", url)
        return None
    except Exception as e:
        logger.warning("Ошибка карточки %s: %s", url, e)
        return None


async def scrape_one_seller(context, seller_url: str, seller_name: str, progress_callback=None) -> List[Dict]:
    seller_page = await context.new_page()
    seller_page.set_default_timeout(15000)

    try:
        if progress_callback:
            await progress_callback(f"{seller_name}: открываю страницу продавца...")

        await seller_page.goto(seller_url, wait_until="domcontentloaded", timeout=30000)
        await seller_page.wait_for_timeout(4000)

        product_links = await collect_product_links(seller_page, LIMIT)

        if not product_links:
            raise RuntimeError(f"{seller_name}: не удалось собрать ссылки на товары")

        if progress_callback:
            await progress_callback(f"{seller_name}: найдено ссылок {len(product_links)}. Читаю карточки...")

        results = []

        for index, url in enumerate(product_links[:LIMIT], start=1):
            item = await safe_parse_product_page(context, url, index)

            if item:
                results.append(item)

            if progress_callback:
                await progress_callback(f"{seller_name}: обработано {index}/{min(len(product_links), LIMIT)}")

            await asyncio.sleep(0.4)

        return results

    finally:
        await seller_page.close()


async def scrape_all_sellers(progress_callback=None) -> Dict[str, List[Dict]]:
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-blink-features=AutomationControlled",
            ],
        )

        context = await browser.new_context(
            viewport={"width": 1440, "height": 2200},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/123.0.0.0 Safari/537.36"
            ),
            locale="ru-RU",
        )

        try:
            result = {}

            for seller in SELLERS:
                items = await scrape_one_seller(
                    context,
                    seller["url"],
                    seller["name"],
                    progress_callback=progress_callback
                )
                result[seller["name"]] = items

            return result

        finally:
            await context.close()
            await browser.close()


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Напиши /novinki и я пришлю Excel-файл с листами МС и ЦР")


async def novinki(update: Update, context: ContextTypes.DEFAULT_TYPE):
    status_msg = await update.message.reply_text("Собираю товары...")

    async def progress(text: str):
        try:
            await status_msg.edit_text(text)
        except Exception:
            pass

    try:
        sellers_data = await asyncio.wait_for(scrape_all_sellers(progress_callback=progress), timeout=600)

        total_items = sum(len(items) for items in sellers_data.values())
        if total_items == 0:
            await status_msg.edit_text("Не удалось собрать данные по товарам.")
            return

        await status_msg.edit_text("Формирую Excel...")

        file_path = save_to_xlsx(sellers_data, OUTPUT_XLSX)

        with open(file_path, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename="wb_products.xlsx",
                caption=(
                    f"Готово:\n"
                    f"МС — {len(sellers_data.get('МС', []))} товаров\n"
                    f"ЦР — {len(sellers_data.get('ЦР', []))} товаров"
                )
            )

        await status_msg.edit_text("Готово.")

    except asyncio.TimeoutError:
        logger.exception("Общий таймаут парсинга")
        await status_msg.edit_text("Ошибка: общий таймаут парсинга.")
    except Exception as e:
        logger.exception("Ошибка")
        await status_msg.edit_text(f"Ошибка: {e}")


async def handle_mentions(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.text:
        return

    text = update.message.text.strip().lower()

    trigger_patterns = [
        f"@{BOT_USERNAME} /novinki",
        f"@{BOT_USERNAME} novinki",
        f"@{BOT_USERNAME} новинки",
        f"/novinki@{BOT_USERNAME}",
    ]

    if any(pattern in text for pattern in trigger_patterns):
        await novinki(update, context)


def main():
    if not TG_TOKEN:
        raise RuntimeError("Добавь TG_BOT_TOKEN в Railway Variables")

    app = Application.builder().token(TG_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("novinki", novinki))
    app.add_handler(MessageHandler(filters.TEXT, handle_mentions))

    logger.info("Bot started")
    app.run_polling()


if __name__ == "__main__":
    main()